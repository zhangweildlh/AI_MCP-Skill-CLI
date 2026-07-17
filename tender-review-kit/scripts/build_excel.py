#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""build_excel.py — 把工作区各专项 md 的清单表格汇总成多 sheet Excel

读一组 md 文件，每个 `## 标题` 下的标准表格 → 一个 Excel sheet。
- 表头按内容自动配色（废标红/评分绿/标识橙/证明紫/时间蓝/矛盾黄）
- 「废标」/「评分」sheet 自动追加"核对结果·责任人"列（方便分配工作 + 打勾）
- 自动跳过非清单 section（如"实质性要求范围定位"这种说明文字）

依赖：openpyxl。用法：
    python build_excel.py <out.xlsx> <md1> [md2 ...]
"""
import sys
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

_SEP = re.compile(r"^\|[\s\-:|]+\|?$")

# 这些关键词的 ## section 是说明性文字,不是清单,不出 sheet
SKIP_KEYWORDS = ["范围定位", "前置判断", "小结", "说明", "导览", "总览", "概况"]
# 这些 section 自动追加"核对·责任人"列
WORKFLOW_KEYWORDS = ["废标", "评分", "证明", "标识"]
# 子节(###)含这些关键词时跳过表格(说明性/台账/自检性子节,非主清单)
SKIP_SUBHEADER_KW = ["台账", "处置", "核验", "对照", "发现", "建议", "边界", "排除", "锚点"]


def parse_sections(md_path):
    """→ [(title, rows)],rows[0] 为表头。
    `## ` 开启/切换 section;`### ` 默认保持 section 内,但跳过说明性子节;`# ` 一级关闭。"""
    secs, title, rows = [], None, []
    in_table = True
    with open(md_path, encoding="utf-8") as f:
        for raw in f:
            s = raw.rstrip("\n").strip()
            if s.startswith("## "):
                if title and rows:
                    secs.append((title, rows))
                title, rows = s[3:].strip(), []
                in_table = True
                continue
            if s.startswith("# ") and not s.startswith("## "):
                if title and rows:
                    secs.append((title, rows))
                title, rows = None, []
                continue
            if s.startswith("### "):
                in_table = not any(kw in s for kw in SKIP_SUBHEADER_KW)
                continue
            if title and in_table and s.startswith("|"):
                if _SEP.match(s):
                    continue
                rows.append([c.strip() for c in s.strip("|").split("|")])
    if title and rows:
        secs.append((title, rows))
    return secs


def header_color(title):
    if "废标" in title:
        return "C00000"
    if "评分" in title:
        return "00B050"
    if "标识" in title or "▲" in title:
        return "ED7D31"
    if "证明" in title or "材料" in title:
        return "7030A0"
    if "时间" in title or "节点" in title:
        return "4472C4"
    if "合同" in title:  # 合同条款·要点(中标后约束,淡黄,提示参考用)
        return "BF8F00"
    if "矛盾" in title or "复核" in title:
        return "BF8F00"
    return "305496"


def safe_name(name, used):
    n = re.sub(r"[\\/?*\[\]:]", "", name)[:28] or "sheet"
    base, i = n, 1
    while n in used:
        n = (base[:26] + str(i))
        i += 1
    used.add(n)
    return n


def should_skip(title):
    return any(kw in title for kw in SKIP_KEYWORDS)


def needs_workflow_cols(title):
    return any(kw in title for kw in WORKFLOW_KEYWORDS)


def main():
    if len(sys.argv) < 3:
        print("用法: python build_excel.py <out.xlsx> <md1> [md2 ...]")
        sys.exit(1)
    out, files = sys.argv[1], sys.argv[2:]
    wb = Workbook()
    wb.remove(wb.active)
    used = set()
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    skipped, made = [], []

    for f in files:
        for title, rows in parse_sections(f):
            if not rows:
                continue
            if should_skip(title):
                skipped.append(title)
                continue

            # 追加"核对结果·责任人·备注"列（仅对工作流 section）
            if needs_workflow_cols(title) and rows:
                add_cols = ["核对结果", "责任人", "完成备注"]
                rows = [r[:] for r in rows]  # 拷贝防污染
                # 表头追加（避免重复）
                if not all(c in rows[0] for c in add_cols):
                    for c in add_cols:
                        if c not in rows[0]:
                            rows[0].append(c)
                    # 数据行补空
                    target_len = len(rows[0])
                    for i in range(1, len(rows)):
                        while len(rows[i]) < target_len:
                            rows[i].append("")

            ws = wb.create_sheet(safe_name(title, used))
            made.append(title)
            fill = PatternFill("solid", fgColor=header_color(title))
            ncol = max(len(r) for r in rows)
            for r, row in enumerate(rows, 1):
                for c in range(1, ncol + 1):
                    val = row[c - 1] if c - 1 < len(row) else ""
                    cell = ws.cell(r, c, val)
                    cell.border = border
                    cell.alignment = Alignment(vertical="center", wrap_text=True)
                    if r == 1:
                        cell.fill = fill
                        cell.font = Font(bold=True, color="FFFFFF")
            ws.freeze_panes = "A2"
            ws.row_dimensions[1].height = 22
            for c in range(1, ncol + 1):
                w = max((len(str(rows[ri][c - 1])) for ri in range(len(rows))
                         if c - 1 < len(rows[ri])), default=8)
                ws.column_dimensions[get_column_letter(c)].width = min(max(w * 1.7, 9), 60)
            # 自动开启表格筛选
            ws.auto_filter.ref = "A1:" + get_column_letter(ncol) + str(len(rows))

    wb.save(out)
    print("OK saved:", out)
    print("sheets (%d):" % len(made), " | ".join(wb.sheetnames))
    if skipped:
        print("已跳过非清单 section (%d):" % len(skipped), " | ".join(skipped))


if __name__ == "__main__":
    main()
