# -*- coding: utf-8 -*-
"""test_qa_full.py — 全脚本回归 + 边界 + 对抗测试(测试员套件)

比 test_smoke.py 更全:覆盖 7 个脚本的正常路径 / 边界 / 对抗输入 + 端到端串联。
**完全自包含**——只依赖 tests/fixtures/sample_tender.docx(合成样本),所有临时件写进
系统临时目录,跑完自清。不碰任何真实标书、不依赖 workspace 状态。

直接跑:  python tests/test_qa_full.py     (exit 0 = 全过,1 = 有失败)
"""
import sys, json, subprocess, tempfile, shutil
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
except Exception:
    pass

KIT = Path(__file__).resolve().parent.parent
SC = KIT / "scripts"
FIX = KIT / "tests" / "fixtures" / "sample_tender.docx"
PY = sys.executable
TMP = Path(tempfile.mkdtemp(prefix="brk_qa_"))

results = []
def ok(name, cond, detail=""):
    results.append((name, bool(cond)))
    print(("[PASS] " if cond else "[FAIL] ") + name + (("  -> " + detail) if (detail and not cond) else ""))

def run(args, **kw):
    return subprocess.run([PY] + [str(a) for a in args], capture_output=True, text=True,
                          encoding="utf-8", errors="replace", **kw)

def write(name, txt):
    p = TMP / name; p.write_text(txt, encoding="utf-8"); return p

def lines_file(name, rows):
    return write(name, "".join("%d\t%s\n" % (n, t) for n, t in rows))

try:
    # ============ extract_text.py ============
    r = run([SC/"extract_text.py", FIX, "--outdir", TMP, "--name", "qa_x"])
    ok("extract: docx 正常退出", r.returncode == 0, r.stderr[:200])
    ok("extract: 生成 lines.txt", (TMP/"qa_x.lines.txt").exists())
    ok("extract: 生成 tables.json(fixture 含表格)", (TMP/"qa_x.tables.json").exists())
    ok("extract: 缺文件 exit=1", run([SC/"extract_text.py", TMP/"nope.docx", "--outdir", TMP]).returncode == 1)
    ok("extract: .doc 老格式 exit=2", run([SC/"extract_text.py", write("x.doc","x"), "--outdir", TMP]).returncode == 2)
    ok("extract: 未知格式 exit=2", run([SC/"extract_text.py", write("x.txt","x"), "--outdir", TMP]).returncode == 2)
    # 空内容文档(模拟扫描件/图片型 PDF 提取为空)→ 必须非零退出,不能静默出空报告
    from docx import Document as _Doc
    _ed = TMP/"empty.docx"; _Doc().save(str(_ed))
    _r = run([SC/"extract_text.py", _ed, "--outdir", TMP])
    ok("extract: 空内容文档 exit!=0(防扫描件静默)", _r.returncode != 0)
    ok("extract: 失败时不留垃圾 lines.txt(先检查后写)", not (TMP/"empty.lines.txt").exists())

    # ============ scan_keywords.py ============
    lf = TMP/"qa_x.lines.txt"
    run([SC/"scan_keywords.py", lf, "--out", TMP/"qa_x.hits.json"])
    H = json.load(open(TMP/"qa_x.hits.json", encoding="utf-8")); s = H["summary"]
    ok("scan: fixture 一级判决词≥12(长词优先去噪后)", s["primary"] >= 12, str(s["primary"]))
    ok("scan: fixture 关系门槛≥6", s["customization"] >= 6, str(s["customization"]))
    ok("scan: fixture 证明文件≥5", s["certifications"] >= 5, str(s["certifications"]))
    ok("scan: fixture ▲ 被识别", "▲" in H["detected_emphasis_marks"])
    r = run([SC/"scan_keywords.py", lines_file("empty.lines.txt", []), "--out", TMP/"empty.hits.json"])
    ok("scan: 空文件不崩 exit=0", r.returncode == 0, r.stderr[:160])
    def detected(rows, name):
        run([SC/"scan_keywords.py", lines_file(name+".lines.txt", rows), "--out", TMP/(name+".hits.json")])
        return json.load(open(TMP/(name+".hits.json"), encoding="utf-8"))["detected_emphasis_marks"]
    ok("scan: 单个 ▲ 也识别(阈值=1)", "▲" in detected([(1,"▲ 投标人必须响应")], "t_tri"))
    ok("scan: 单个 ★ 也识别(阈值=1)", "★" in detected([(1,"★ 投标人必须响应")], "t_star"))
    ok("scan: 2 个 ● 项目符号不误报(阈值=3)", "●" not in detected([(1,"● 一"),(2,"● 二")], "t_dot"))
    star_pos = detected([(1,"* 投标人必须A"),(2,"* 投标人必须B"),(3,"[T1] 参数 | * 非首列说明"),(4,"[T2] * 首列参数 | 说明")], "t_starpos")
    ok("scan: * 强调(行首/单元格/表格首列)达阈值被识别", "*" in star_pos, str(star_pos))
    ok("scan: * 少于3次按设计不识别", "*" not in detected([(1,"* A"),(2,"* B")], "t_starlow"))
    ok("scan: * 乘号/加粗/脚注不误报", "*" not in detected([(1,"算 3*5"),(2,"**加粗**"),(3,"见备注* 后文")], "t_starneg"))
    run([SC/"scan_keywords.py", lines_file("excl.lines.txt", [(1,"第三章 否决与无效情形"),(2,"目 录")]), "--out", TMP/"excl.hits.json"])
    ok("scan: 第X章/目录行被排除", json.load(open(TMP/"excl.hits.json", encoding="utf-8"))["summary"]["primary"] == 0)
    run([SC/"scan_keywords.py", lines_file("toc_body.lines.txt", [(1,"技术参数支持多级目录树，投标无效情形按本条执行")]), "--out", TMP/"toc_body.hits.json"])
    ok("scan: 正文含目录二字不误排除", json.load(open(TMP/"toc_body.hits.json", encoding="utf-8"))["summary"]["primary"] >= 1)
    run([SC/"scan_keywords.py", lines_file("multi_mark.lines.txt", [(1,"[T1] 参数 | ▲1 支持HDMI | ▲2 支持SDI | ★3 提供检测报告")]), "--out", TMP/"multi_mark.hits.json"])
    mh = json.load(open(TMP/"multi_mark.hits.json", encoding="utf-8"))
    ok("scan: 同一表格行多个 ▲/★ 逐条记录", mh["summary"]["emphasis_marks"] == 3, str(mh["hits"]["emphasis_marks"]))
    run([SC/"scan_keywords.py", lines_file("longest.lines.txt", [(1,"供应商按照无效投标处理")]), "--out", TMP/"longest.hits.json"])
    lw = [h["word"] for h in json.load(open(TMP/"longest.hits.json", encoding="utf-8"))["hits"]["primary"]]
    ok("scan: primary 同行重叠命中长词优先", lw == ["按照无效投标处理"], str(lw))

    # ============ scan_candidates.py ============
    kw_before = len(json.load(open(KIT/"data"/"keywords.json", encoding="utf-8"))["categories"][0]["words"])
    run([SC/"scan_candidates.py", lf])
    ok("candidates: 默认写同目录 .candidates.json", (TMP/"qa_x.candidates.json").exists())
    ok("candidates: 不写进 data/", not (KIT/"data"/"candidates.json").exists())
    kw_after = len(json.load(open(KIT/"data"/"keywords.json", encoding="utf-8"))["categories"][0]["words"])
    ok("candidates: 不自动改 keywords.json", kw_before == kw_after)
    cd = json.load(open(TMP/"qa_x.candidates.json", encoding="utf-8")).get("candidates", [])
    ok("candidates: 全部 pending_review", all(c.get("status")=="pending_review" for c in cd) or not cd)

    # ============ check_coverage.py ============
    def cov(hit_line, scope, md_text, extra=None):
        hp = TMP/"cv.hits.json"; json.dump({"hits":{"primary":[{"line":hit_line,"word":"否决","scope":scope,"text":"x"}]}}, open(hp,"w",encoding="utf-8"))
        mp = write("cv.工作区.md", md_text); op = TMP/"cv.coverage.json"
        rr = run([SC/"check_coverage.py", hp, mp, "--out", op] + (extra or []))
        return rr, json.load(open(op, encoding="utf-8"))
    TBL = "## 商务线·废标\n\n| ID | 废标条款 | 出处(行号) |\n|---|---|---|\n%s\n"
    ok("coverage: 引用精确同行->已覆盖", cov(103,["bid_phase"],TBL % "| D1 | x | 行103 |")[1]["uncovered"] == 0)
    ok("coverage: ±0 下行100 不覆盖行103", cov(103,["bid_phase"],TBL % "| D1 | x | 行100 |")[1]["uncovered"] == 1)
    ok("coverage: 范围 行100-105 覆盖行103", cov(103,["bid_phase"],TBL % "| D1 | x | 行100-105 |")[1]["uncovered"] == 0)
    ok("coverage: 第103行 也算覆盖", cov(103,["bid_phase"],TBL % "| D1 | x | 第103行 |")[1]["uncovered"] == 0)
    ok("coverage: 裸数字103 不算覆盖(防金额/数量误判)", cov(103,["bid_phase"],TBL % "| D1 | x | 103 |")[1]["uncovered"] == 1)
    ledger = "## 商务线·废标\n\n| ID | 条款 | 出处 |\n|---|---|---|\n| D1 | x | 行999 |\n\n### 处置台账\n\n| 命中行 | 处置 |\n|---|---|\n| 103 | 已排除 |\n"
    ok("coverage: ### 台账行号不算覆盖(防自证)", cov(103,["bid_phase"],ledger)[1]["uncovered"] == 1)
    ok("coverage: --strict 遇 high 未覆盖 exit!=0", cov(103,["bid_phase"],TBL % "| D1 | x | 行100 |",["--strict"])[0].returncode != 0)
    ok("coverage: 纯 contract_phase 不计 in_scope", cov(103,["contract_phase"],TBL % "| D1 | x | 行100 |")[1]["in_scope_primary"] == 0)
    rr, c = cov(103,["bid_phase"],"## 评分\n\n| 项 | 分 |\n|---|---|\n| a | 1分 |\n")
    ok("coverage: 无废标节也不崩,全未覆盖", rr.returncode == 0 and c["uncovered"] == 1)

    # ============ check_completeness.py ============
    def comp(md_text, hits=None, extra=None):
        a = [SC/"check_completeness.py", write("cp.工作区.md", md_text)]
        if hits: a += ["--hits", hits]
        if extra: a += extra
        return run(a)
    few = "## 商务线·废标\n\n| ID | 条款 | 出处 |\n|---|---|---|\n| D1 | a | 行1 |\n| D2 | b | 行2 |\n"
    ok("completeness: 废标偏少给 warning", "废标条数" in comp(few).stdout)
    ok("completeness: --strict 有 warning exit!=0", comp(few, extra=["--strict"]).returncode != 0)
    r = comp("## 评分\n\n| 评分项 | 梯度 |\n|---|---|\n| 业绩 | 近三年三项以上得满额 |\n| 价格 | 最低价优先 |\n")
    ok("completeness: 评分行缺『分』字被逮", "不含" in r.stdout)
    hp = TMP/"cp.hits.json"; json.dump({"summary":{"emphasis_marks":10}}, open(hp,"w",encoding="utf-8"))
    r = comp("## 技术线·▲标识\n\n| ID | 参数 | 出处 |\n|---|---|---|\n| E1 | a | 行1 |\n| E2 | b | 行2 |\n", hits=str(hp))
    ok("completeness: ▲ 清单远少于撒网->warning", "▲" in r.stdout and "压缩" in r.stdout)

    # ============ harvest_ai_words.py ============
    hv_lines = lines_file("hv.lines.txt", [(1, "供应商出现测试违规的,取消测试资格。")])
    hv_hits = TMP/"hv.hits.json"
    json.dump({"hits":{"primary":[],"contract":[],"secondary":[],"customization":[],"certifications":[],"emphasis_marks":[]}},
              open(hv_hits, "w", encoding="utf-8"), ensure_ascii=False)
    hv_wl = write("hv.工作区.md",
        "# AI 发现回扫\n\n## 商务线·废标\n\n| ID | 条款 | 出处(行号) |\n|---|---|---|\n"
        "| D1 | 已发现条款 | 行1 |\n\n"
        "## AI发现疑似判词\n\n| 疑似判词 | 原文摘要 | 出处 | 建议分类 |\n|---|---|---|---|\n"
        "| 取消测试资格 | 出现测试违规的,取消测试资格 | 行1 | primary/bid_phase |\n")
    r = run([SC/"harvest_ai_words.py", hv_wl, "--lines", hv_lines, "--hits", hv_hits])
    ok("harvest: 默认模式会临时回扫当前标书", r.returncode == 0 and "临时回扫当前标书" in r.stdout)
    ok("harvest: 未入库也生成当前补漏 hits", (TMP/"hv.ai_rescan.hits.json").exists())
    hv_rescan = json.load(open(TMP/"hv.ai_rescan.hits.json", encoding="utf-8"))
    ok("harvest: 临时回扫命中新 AI 判词", any(h.get("word") == "取消测试资格" for h in hv_rescan["hits"]["primary"]))
    r = run([SC/"harvest_ai_words.py", hv_wl, "--reject-all"])
    ok("harvest: 拒绝入库只清 pending,不代表忽略当前补漏", r.returncode == 0 and "当前标书补漏不受影响" in r.stdout)

    # ============ cross_doc.py ============
    ok("cross_doc: <2 文件 exit=1", run([SC/"cross_doc.py", lf]).returncode == 1)
    a = lines_file("d1.lines.txt", [(1,"最高投标限价为 100 万元")]); b = lines_file("d2.lines.txt", [(1,"最高投标限价为 200 万元")])
    ok("cross_doc: 限价 100万 vs 200万 报矛盾", "金额矛盾" in run([SC/"cross_doc.py", a, b]).stdout)
    a2 = lines_file("d3.lines.txt", [(1,"投标保证金 100 万元")]); b2 = lines_file("d4.lines.txt", [(1,"投标保证金 1000000 元")])
    out = run([SC/"cross_doc.py", a2, b2]).stdout
    ok("cross_doc: 100万元==1000000元 不误报(单位归一)", "未发现确定性矛盾" in out or "矛盾数：0" in out)
    a3 = lines_file("d5.lines.txt", [(1,"投标截止 2024年8月22日")]); b3 = lines_file("d6.lines.txt", [(1,"投标截止 2024年9月1日")])
    ok("cross_doc: 投标截止日期不一致报矛盾", "日期矛盾" in run([SC/"cross_doc.py", a3, b3]).stdout)

    # ============ build_excel.py(用合成工作区,不依赖真实标书) ============
    ok("build_excel: 缺参数 exit=1", run([SC/"build_excel.py", TMP/"x.xlsx"]).returncode == 1)
    synth_wl = write("synth.工作区.md",
        "# QA 合成工作区\n\n## 商务线·废标\n\n| ID | 类别 | 出处(行号) |\n|---|---|---|\n"
        "| D1 | 资格性 | 行10 |\n| D2 | 文件性 | 行14 |\n| D3 | 商务实质 | 行16 |\n\n"
        "### 撒网处置台账\n\n| 命中行 | 处置 |\n|---|---|\n| 99 | 已排除 |\n| 98 | 已排除 |\n\n"
        "## 技术线·▲标识\n\n| ID | 标识 | 参数 | 出处 |\n|---|---|---|---|\n| E1 | ▲ | 频率≥2.4G | 行27 |\n")
    xlsx = TMP/"qa_out.xlsx"
    r = run([SC/"build_excel.py", xlsx, synth_wl])
    ok("build_excel: 合成工作区出 Excel 成功", r.returncode == 0 and xlsx.exists(), r.stderr[:200])
    if xlsx.exists():
        from openpyxl import load_workbook
        wb = load_workbook(xlsx)
        disq = [n for n in wb.sheetnames if "废标" in n]
        ok("build_excel: 有废标 sheet", len(disq) >= 1, str(wb.sheetnames))
        if disq:
            w = wb[disq[0]]
            ok("build_excel: 废标表=表头+3行(台账2行被排除)", w.max_row == 4, "max_row=%d" % w.max_row)
            ok("build_excel: 废标 sheet 追加核对/责任人列", "核对结果" in [c.value for c in w[1]])
        ok("build_excel: 有▲标识 sheet", any("标识" in n for n in wb.sheetnames), str(wb.sheetnames))

    # ============ run_pipeline.py verify fail-fast ============
    bad_wl = write("bad.工作区.md",
        "# 失败护栏复现\n\n## 商务线·废标\n\n| ID | 条款 | 出处(行号) |\n|---|---|---|\n| D1 | 测试 | 行1 |\n")
    write("bad.hits.json", "{ not json")
    bad_xlsx = TMP/"bad.xlsx"
    rr = run([KIT/"run_pipeline.py", "verify", bad_wl])
    ok("run_pipeline: 护栏脚本失败时 verify 非零退出", rr.returncode != 0)
    ok("run_pipeline: 护栏失败时不生成 Excel", not bad_xlsx.exists())

    # ============ promote_candidates.py(用合成候选文件) ============
    syn_cand = write("syn.candidates.json", json.dumps({"candidates":[
        {"word":"视为弃标A","occurrences":3,"contexts":["x"],"suggested_scope":["bid_phase"],"suggested_category":"primary","status":"pending_review","near_known":True}
    ]}, ensure_ascii=False))
    r = run([SC/"promote_candidates.py", "--list", "--candidates", syn_cand])
    ok("promote: --list 跑通(指定候选文件)", r.returncode == 0, (r.stderr or r.stdout)[:160])

    # ============ 端到端串联 ============
    bidlines = [h["line"] for h in H["hits"]["primary"] if "bid_phase" in h.get("scope",[])][:10]
    rowtxt = "".join("| D%02d | 合规项 | 行%d |\n" % (i+1, ln) for i, ln in enumerate(bidlines))
    e2e = write("e2e.工作区.md",
        "# 端到端\n\n## 商务线·废标\n\n| ID | 类别 | 出处(行号) |\n|---|---|---|\n" + rowtxt +
        "\n## 商务线·评分\n\n| 评分项 | 梯度 |\n|---|---|\n| 价格分 | 最低价得40分 |\n| 业绩分 | 每项2分 |\n")
    rr = run([SC/"check_coverage.py", TMP/"qa_x.hits.json", e2e, "--out", TMP/"e2e.coverage.json"])
    e2ecov = json.load(open(TMP/"e2e.coverage.json", encoding="utf-8"))
    ok("E2E: extract→scan→coverage 跑通", rr.returncode == 0)
    ok("E2E: 引用的 bid 行基本被覆盖", e2ecov["covered"] >= len(set(bidlines)) - 2, "covered=%d" % e2ecov["covered"])
    ok("E2E: completeness 跑通", run([SC/"check_completeness.py", e2e, "--hits", TMP/"qa_x.hits.json"]).returncode == 0)
    rr = run([SC/"build_excel.py", TMP/"e2e.xlsx", e2e])
    ok("E2E: build_excel 跑通且产出文件", rr.returncode == 0 and (TMP/"e2e.xlsx").exists())

finally:
    shutil.rmtree(TMP, ignore_errors=True)

npass = sum(1 for _, c in results if c); nfail = len(results) - npass
print("\n" + "=" * 60)
print("总账: %d 项,%d PASS,%d FAIL" % (len(results), npass, nfail))
if nfail:
    print("失败项:")
    for n, c in results:
        if not c: print("  [FAIL]", n)
sys.exit(1 if nfail else 0)
